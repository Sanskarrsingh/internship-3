from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import sqlite3
from datetime import datetime, time, timedelta, date
import pytz
import uuid
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from dotenv import load_dotenv
import json
import pdfkit
import base64
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

app = Flask(__name__, static_folder='frontend/build')
app.secret_key = 'your_secret_key'
CORS(app, resources={r"/*": {"origins": "*"}})

# Load environment variables
load_dotenv()

SMTP_SERVER = os.getenv('SMTP_SERVER')
SMTP_PORT = int(os.getenv('SMTP_PORT', 587))
SMTP_USERNAME = os.getenv('SMTP_USERNAME')
SMTP_PASSWORD = os.getenv('SMTP_PASSWORD')
ADMIN_EMAIL = os.getenv('ADMIN_EMAIL')

# Path to wkhtmltopdf
WKHTMLTOPDF_PATH = os.getenv('WKHTMLTOPDF_PATH', '/usr/local/bin/wkhtmltopdf')

# Configurable time range for task submission
SUBMISSION_START_TIME = time(9, 0)  # 9:00 AM IST
SUBMISSION_END_TIME = time(19, 30)  # 7:30 PM IST
IST = pytz.timezone('Asia/Kolkata')

pdfkit_config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

def init_db():
    conn = sqlite3.connect('ted.db')
    cursor = conn.cursor()

    # TED system tables
    cursor.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id TEXT NOT NULL UNIQUE,
        email TEXT NOT NULL,
        password TEXT NOT NULL,
        role TEXT NOT NULL,
        status TEXT DEFAULT 'active'
    )''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS tasks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id TEXT NOT NULL,
        area_of_effort TEXT,
        effort_hours INTEGER,
        effort_minutes INTEGER,
        effort_towards TEXT,
        time_log_type TEXT,
        manager_note TEXT,
        broad_area_of_work TEXT,
        reviewer_note TEXT,
        output_file TEXT,
        output_location TEXT,
        task_date DATE DEFAULT (DATE('now', 'localtime'))
    )''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS invitations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT NOT NULL UNIQUE,
        user_id TEXT NOT NULL,
        role TEXT NOT NULL,
        invitation_code TEXT NOT NULL,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    # Payroll system tables
    cursor.execute('''CREATE TABLE IF NOT EXISTS employees (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        email TEXT NOT NULL UNIQUE,
        department TEXT NOT NULL
    )''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS payroll (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        employee_id INTEGER NOT NULL,
        period TEXT NOT NULL,
        salary REAL NOT NULL,
        bonus REAL,
        deductions REAL,
        tax REAL,
        net_salary REAL NOT NULL,
        FOREIGN KEY (employee_id) REFERENCES employees(id)
    )''')
    
    conn.commit()
    conn.close()

init_db()

def is_within_submission_time():
    now = datetime.now(IST).time()
    return SUBMISSION_START_TIME <= now <= SUBMISSION_END_TIME

def count_weekdays(start_date, end_date):
    weekdays = 0
    current_date = start_date
    while current_date <= end_date:
        if current_date.weekday() < 5:  # Monday to Friday are 0-4
            weekdays += 1
        current_date += timedelta(days=1)
    return weekdays

def send_invitation_email(email, invitation_code):
    try:
        msg = MIMEText(f'You have been invited to register. Use the following invitation code to register: {invitation_code}')
        msg['Subject'] = 'Invitation to Register'
        msg['From'] = SMTP_USERNAME
        msg['To'] = email

        print(f"Connecting to SMTP server: {SMTP_SERVER}:{SMTP_PORT}")
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.ehlo()  # Can be omitted
            server.starttls()
            server.ehlo()  # Can be omitted
            print("Logging in to SMTP server")
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            print(f"Sending email to {email}")
            server.sendmail(SMTP_USERNAME, email, msg.as_string())
        print(f"Email sent to {email}")
    except Exception as e:
        print(f"Failed to send email: {e}")
        raise e  # Re-raise the exception to be handled by the caller

# TED system routes (Existing routes)

# Invite route
@app.route('/invite', methods=['POST'])
def invite():
    data = request.json
    print(f"Received data: {data}")  # Log incoming data

    email = data.get('email')
    user_id = data.get('user_id')
    role = data.get('role')
    
    if not email or not role or not user_id:
        return jsonify({'message': 'Email, user_id, and role are required'}), 400

    invitation_code = str(uuid.uuid4())  # Generate a unique invitation code

    if role not in ['employee', 'manager', 'reviewer']:
        return jsonify({'message': 'Invalid role'}), 400

    conn = sqlite3.connect('ted.db')
    cursor = conn.cursor()
    try:
        # Check if an invitation already exists for this email
        cursor.execute('SELECT * FROM invitations WHERE email = ? OR user_id = ?', (email, user_id))
        existing_invitation = cursor.fetchone()
        if existing_invitation:
            print("Invitation already exists for this email or user_id")
            return jsonify({'message': 'Invitation already exists for this email or user_id'}), 400
        
        print(f"Attempting to insert invitation into database for {email} with role {role} and code {invitation_code}")
        cursor.execute('INSERT INTO invitations (email, user_id, role, invitation_code) VALUES (?, ?, ?, ?)', (email, user_id, role, invitation_code))
        conn.commit()
        print("Invitation inserted into database successfully")
        send_invitation_email(email, invitation_code)
        print("Invitation email sent successfully")
    except Exception as e:
        print(f"Error: {e}")  # Log any other errors
        return jsonify({'message': f'An error occurred: {e}'}), 500
    finally:
        conn.close()
    return jsonify({'message': 'Invitation sent successfully'}), 201

# Register route
@app.route('/register', methods=['POST'])
def register():
    data = request.json
    user_id = data['user_id']
    email = data['email']
    password = data['password']
    role = data['role']
    invitation_code = data.get('invitation')

    conn = sqlite3.connect('ted.db')
    cursor = conn.cursor()

    # Fetch special users from the JSON file
    with open('special_users.json', 'r') as f:
        special_users = json.load(f)

    print(f"Registering user: {user_id}, email: {email}, role: {role}, invitation: {invitation_code}")

    # Check if the user is in the special users list
    if (role == 'manager' and user_id in special_users['managers']) or (role == 'reviewer' and user_id in special_users['reviewers']):
        try:
            cursor.execute('INSERT INTO users (user_id, email, password, role) VALUES (?, ?, ?, ?)', (user_id, email, password, role))
            conn.commit()
            print("Special user registered successfully")
        except sqlite3.IntegrityError:
            print("User ID or email already exists")
            return jsonify({'message': 'User ID or email already exists'}), 400
        finally:
            conn.close()
        return jsonify({'user_id': user_id, 'role': role}), 201

    # Otherwise, check the invitation code and user_id
    cursor.execute('SELECT * FROM invitations WHERE user_id = ? AND role = ? AND invitation_code = ?', (user_id, role, invitation_code))
    invitation = cursor.fetchone()
    if not invitation:
        print("Invalid invitation or role")
        return jsonify({'message': 'Invalid invitation or role'}), 400

    try:
        cursor.execute('INSERT INTO users (user_id, email, password, role) VALUES (?, ?, ?, ?)', (user_id, email, password, role))
        conn.commit()
        cursor.execute('DELETE FROM invitations WHERE user_id = ? AND role = ? AND invitation_code = ?', (user_id, role, invitation_code))
        conn.commit()
        print("User registered successfully")
    except sqlite3.IntegrityError:
        print("User ID or email already exists")
        return jsonify({'message': 'User ID or email already exists'}), 400
    finally:
        conn.close()
    return jsonify({'user_id': user_id, 'role': role}), 201

# Login route
@app.route('/login', methods=['POST'])
def login():
    data = request.json
    user_id = data['user_id']
    password = data['password']
    
    print(f"Login attempt for user_id: {user_id}")

    conn = sqlite3.connect('ted.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users WHERE user_id = ? AND password = ?', (user_id, password))
    user = cursor.fetchone()
    conn.close()
    if user:
        print("Login successful")
        return jsonify({'user_id': user[1], 'role': user[4]}), 200
    else:
        print("Login failed: Invalid credentials")
        return jsonify({'message': 'Invalid credentials'}), 401

# Add task route
@app.route('/add_task', methods=['POST'])
def add_task():
    if not is_within_submission_time():
        return jsonify({'message': 'Tasks can only be submitted between 9:00 AM and 7:30 PM IST'}), 403

    data = request.json
    user_id = data['user_id']
    area_of_effort = data['area_of_effort']
    effort_hours = int(data['effort_hours']) if data['effort_hours'] else 0
    effort_minutes = int(data['effort_minutes']) if data['effort_minutes'] else 0
    effort_towards = data['effort_towards']
    time_log_type = data['time_log_type']
    output_file = data['output_file']
    output_location = data['output_location']
    task_date = data.get('task_date', None)

    # Check if the task_date is not a previous date
    if task_date:
        task_date_obj = datetime.strptime(task_date, '%Y-%m-%d').date()
        today = date.today()
        if task_date_obj < today:
            return jsonify({'message': 'Cannot add a task for a previous date'}), 400

    if effort_hours < 0 or effort_minutes < 0:
        return jsonify({'message': 'Effort hours and minutes must be non-negative'}), 400

    conn = sqlite3.connect('ted.db')
    cursor = conn.cursor()
    cursor.execute('SELECT role FROM users WHERE user_id = ?', (user_id,))
    user_role = cursor.fetchone()
    if user_role and user_role[0] != 'employee':
        return jsonify({'message': 'Only employees can add tasks'}), 403

    cursor.execute('''INSERT INTO tasks 
        (user_id, area_of_effort, effort_hours, effort_minutes, effort_towards, time_log_type, output_file, output_location, task_date)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', 
        (user_id, area_of_effort, effort_hours, effort_minutes, effort_towards, time_log_type, output_file, output_location, task_date))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Task added successfully'}), 201

# Update task route
@app.route('/update_task', methods=['PUT'])
def update_task():
    data = request.json
    task_id = data['task_id']
    area_of_effort = data['area_of_effort']
    effort_hours = int(data['effort_hours']) if data['effort_hours'] else 0
    effort_minutes = int(data['effort_minutes']) if data['effort_minutes'] else 0
    effort_towards = data['effort_towards']
    time_log_type = data['time_log_type']
    output_file = data['output_file']
    output_location = data['output_location']
    
    conn = sqlite3.connect('ted.db')
    cursor = conn.cursor()

    cursor.execute('SELECT task_date FROM tasks WHERE id = ?', (task_id,))
    task_date = cursor.fetchone()
    if task_date:
        task_date_str = task_date[0]
        task_date = datetime.strptime(task_date_str, '%Y-%m-%d').date()
        today = datetime.now().date()
        if task_date != today:
            conn.close()
            return jsonify({'message': 'You can only edit tasks for the current day'}), 403

    cursor.execute('''
        UPDATE tasks SET area_of_effort = ?, effort_hours = ?, effort_minutes = ?, effort_towards = ?, time_log_type = ?, output_file = ?, output_location = ?
        WHERE id = ?
    ''', (area_of_effort, effort_hours, effort_minutes, effort_towards, time_log_type, output_file, output_location, task_id))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Task updated successfully'}), 200

# Get tasks for a specific date
@app.route('/get_tasks_for_date', methods=['GET'])
def get_tasks_for_date():
    user_id = request.args.get('user_id')
    task_date = request.args.get('task_date')
    conn = sqlite3.connect('ted.db')
    cursor = conn.cursor()
    if user_id:
        cursor.execute('SELECT * FROM tasks WHERE user_id = ? AND task_date = ?', (user_id, task_date))
    else:
        cursor.execute('SELECT * FROM tasks WHERE task_date = ?', (task_date,))
    tasks = cursor.fetchall()
    conn.close()
    task_list = [
        {
            'id': task[0],
            'user_id': task[1],
            'area_of_effort': task[2],
            'effort_hours': task[3],
            'effort_minutes': task[4],
            'effort_towards': task[5],
            'time_log_type': task[6],
            'manager_note': task[7],
            'broad_area_of_work': task[8],
            'reviewer_note': task[9],
            'output_file': task[10],
            'output_location': task[11],
            'task_date': task[12]
        } for task in tasks
    ]
    total_effort_hours = sum(int(task['effort_hours'] or 0) + int(task['effort_minutes'] or 0) / 60 for task in task_list)
    return jsonify({'tasks': task_list, 'total_effort_hours': total_effort_hours}), 200

# Get tasks for a period
@app.route('/get_tasks_for_period', methods=['GET'])
def get_tasks_for_period():
    user_id = request.args.get('user_id')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    conn = sqlite3.connect('ted.db')
    cursor = conn.cursor()
    query = 'SELECT * FROM tasks WHERE task_date BETWEEN ? AND ?'
    params = [from_date, to_date]
    if user_id:
        query += ' AND user_id = ?'
        params.append(user_id)
    cursor.execute(query, params)
    tasks = cursor.fetchall()
    conn.close()
    task_list = [
        {
            'id': task[0],
            'user_id': task[1],
            'area_of_effort': task[2],
            'effort_hours': task[3],
            'effort_minutes': task[4],
            'effort_towards': task[5],
            'time_log_type': task[6],
            'manager_note': task[7],
            'broad_area_of_work': task[8],
            'reviewer_note': task[9],
            'output_file': task[10],
            'output_location': task[11],
            'task_date': task[12]
        } for task in tasks
    ]
    return jsonify(task_list), 200

# Get report route
@app.route('/get_report', methods=['GET'])
def get_report():
    user_id = request.args.get('user_id')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')

    conn = sqlite3.connect('ted.db')
    cursor = conn.cursor()

    cursor.execute('''
        WITH RECURSIVE dates(date) AS (
            SELECT ?
            UNION ALL
            SELECT date(date, '+1 day')
            FROM dates
            WHERE date < ?
        )
        SELECT date
        FROM dates
        WHERE date NOT IN (
            SELECT task_date
            FROM tasks
            WHERE user_id = ?
            AND task_date BETWEEN ? AND ?
        )
        AND strftime('%w', date) NOT IN ('0', '6')
    ''', (from_date, to_date, user_id, from_date, to_date))
    missed_dates = [row[0] for row in cursor.fetchall()]

    cursor.execute('''
        SELECT SUM(effort_hours + effort_minutes / 60.0) AS total_hours
        FROM tasks
        WHERE user_id = ? AND task_date BETWEEN ? AND ?
    ''', (user_id, from_date, to_date))
    total_effort_hours = cursor.fetchone()[0] or 0.0

    cursor.execute('''
        SELECT COALESCE(NULLIF(broad_area_of_work, ''), 'Undefined') AS broad_area_of_work,
               SUM(effort_hours + effort_minutes / 60.0) AS total_hours
        FROM tasks
        WHERE user_id = ? AND task_date BETWEEN ? AND ?
        GROUP BY broad_area_of_work
    ''', (user_id, from_date, to_date))
    broad_area_of_work_hours = cursor.fetchall()

    cursor.execute('''
        SELECT task_date
        FROM tasks
        WHERE user_id = ? AND task_date BETWEEN ? AND ?
        GROUP BY task_date
        HAVING SUM(effort_hours + effort_minutes / 60.0) < 8
    ''', (user_id, from_date, to_date))
    less_than_8_hours_dates = [row[0] for row in cursor.fetchall()]

    cursor.execute('''
        SELECT task_date, area_of_effort
        FROM tasks
        WHERE user_id = ? AND task_date BETWEEN ? AND ? AND (output_file IS NULL OR output_file = '')
    ''', (user_id, from_date, to_date))
    missing_files_tasks = cursor.fetchall()

    cursor.execute('''
        SELECT effort_towards, SUM(effort_hours + effort_minutes / 60.0) AS total_hours
        FROM tasks
        WHERE user_id = ? AND task_date BETWEEN ? AND ?
        GROUP BY effort_towards
    ''', (user_id, from_date, to_date))
    effort_towards_hours = cursor.fetchall()

    cursor.execute('''
        SELECT task_date, area_of_effort, manager_note, reviewer_note
        FROM tasks
        WHERE user_id = ? AND task_date BETWEEN ? AND ?
    ''', (user_id, from_date, to_date))
    notes = cursor.fetchall()

    start_date = datetime.strptime(from_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(to_date, '%Y-%m-%d').date()
    total_working_days = count_weekdays(start_date, end_date)

    conn.close()

    report = {
        'missed_dates': missed_dates,
        'total_effort_hours': float(total_effort_hours),
        'broad_area_of_work_hours': [(item[0], float(item[1])) for item in broad_area_of_work_hours],
        'less_than_8_hours_dates': less_than_8_hours_dates,
        'missing_files_tasks': [(task[0], task[1]) for task in missing_files_tasks],
        'effort_towards_hours': [(item[0], float(item[1])) for item in effort_towards_hours],
        'notes': notes,
        'total_working_days': total_working_days
    }

    return jsonify(report), 200

# Delete task route
@app.route('/delete_task', methods=['DELETE'])
def delete_task():
    task_id = request.args.get('task_id')
    conn = sqlite3.connect('ted.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM tasks WHERE id = ?', (task_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Task deleted successfully'}), 200

# Add manager note route
@app.route('/add_manager_note', methods=['POST'])
def add_manager_note():
    data = request.json
    task_id = data['task_id']
    manager_note = data['manager_note']
    broad_area_of_work = data['broad_area_of_work']
    conn = sqlite3.connect('ted.db')
    cursor = conn.cursor()
    cursor.execute('UPDATE tasks SET manager_note = ?, broad_area_of_work = ? WHERE id = ?', 
        (manager_note, broad_area_of_work, task_id))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Manager note added successfully'}), 200

# Add reviewer note route
@app.route('/add_reviewer_note', methods=['POST'])
def add_reviewer_note():
    data = request.json
    task_id = data['task_id']
    reviewer_note = data['reviewer_note']
    conn = sqlite3.connect('ted.db')
    cursor = conn.cursor()
    cursor.execute('UPDATE tasks SET reviewer_note = ? WHERE id = ?', 
        (reviewer_note, task_id))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Reviewer note added successfully'}), 200

# Get users route
@app.route('/get_users', methods=['GET'])
def get_users():
    conn = sqlite3.connect('ted.db')
    cursor = conn.cursor()
    cursor.execute("SELECT user_id, email FROM users WHERE role = 'employee'")
    users = cursor.fetchall()
    conn.close()
    user_list = [{'user_id': user[0], 'email': user[1]} for user in users]
    return jsonify(user_list), 200

# Special users route
@app.route('/special_users', methods=['GET'])
def special_users():
    return send_from_directory(os.path.dirname(__file__), 'special_users.json')

# Send report route
@app.route('/send_report', methods=['POST'])
def send_report():
    data = request.json
    user_id = data['user_id']
    from_date = data.get('from_date')
    to_date = data.get('to_date')
    task_date = data.get('task_date')
    role = data.get('role')  # Get the role from the request data

    conn = sqlite3.connect('ted.db')
    cursor = conn.cursor()
    
    cursor.execute('SELECT email, role FROM users WHERE user_id = ?', (user_id,))
    user_info = cursor.fetchone()
    if not user_info:
        return jsonify({'message': 'User not found'}), 404

    user_email, user_role = user_info

    if task_date:
        cursor.execute('SELECT * FROM tasks WHERE user_id = ? AND task_date = ?', (user_id, task_date))
    else:
        cursor.execute('SELECT * FROM tasks WHERE user_id = ? AND task_date BETWEEN ? AND ?', (user_id, from_date, to_date))
    tasks = cursor.fetchall()

    if not tasks:
        conn.close()
        return jsonify({'message': 'No tasks found for the specified date(s)'}), 404

    task_list = [
        {
            'id': task[0],
            'user_id': task[1],
            'area_of_effort': task[2],
            'effort_hours': task[3],
            'effort_minutes': task[4],
            'effort_towards': task[5],
            'time_log_type': task[6],
            'manager_note': task[7],
            'broad_area_of_work': task[8],
            'reviewer_note': task[9],
            'output_file': task[10],
            'output_location': task[11],
            'task_date': task[12]
        } for task in tasks
    ]

    tasks_by_date = {}
    for task in task_list:
        if task['task_date'] not in tasks_by_date:
            tasks_by_date[task['task_date']] = []
        tasks_by_date[task['task_date']].append(task)

    total_effort_hours = sum(task['effort_hours'] + task['effort_minutes'] / 60 for task in task_list)

    if not task_date:
        cursor.execute('''
            WITH RECURSIVE dates(date) AS (
                SELECT ?
                UNION ALL
                SELECT date(date, '+1 day')
                FROM dates
                WHERE date < ?
            )
            SELECT date
            FROM dates
            WHERE date NOT IN (
                SELECT task_date
                FROM tasks
                WHERE user_id = ?
                AND task_date BETWEEN ? AND ?
            )
            AND strftime('%w', date) NOT IN ('0', '6')
        ''', (from_date, to_date, user_id, from_date, to_date))
        missed_dates = [row[0] for row in cursor.fetchall()]

        cursor.execute('''
            SELECT COALESCE(NULLIF(broad_area_of_work, ''), 'Undefined') AS broad_area_of_work,
                   SUM(effort_hours + effort_minutes / 60.0) AS total_hours
            FROM tasks
            WHERE user_id = ? AND task_date BETWEEN ? AND ?
            GROUP BY broad_area_of_work
        ''', (user_id, from_date, to_date))
        broad_area_of_work_hours = cursor.fetchall()

        cursor.execute('''
            SELECT task_date
            FROM tasks
            WHERE user_id = ? AND task_date BETWEEN ? AND ?
            GROUP BY task_date
            HAVING SUM(effort_hours + effort_minutes / 60.0) < 8
        ''', (user_id, from_date, to_date))
        less_than_8_hours_dates = [row[0] for row in cursor.fetchall()]

        cursor.execute('''
            SELECT task_date, area_of_effort
            FROM tasks
            WHERE user_id = ? AND task_date BETWEEN ? AND ? AND (output_file IS NULL OR output_file = '')
        ''', (user_id, from_date, to_date))
        missing_files_tasks = cursor.fetchall()

        cursor.execute('''
            SELECT effort_towards, SUM(effort_hours + effort_minutes / 60.0) AS total_hours
            FROM tasks
            WHERE user_id = ? AND task_date BETWEEN ? AND ?
            GROUP BY effort_towards
        ''', (user_id, from_date, to_date))
        effort_towards_hours = cursor.fetchall()

        start_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(to_date, '%Y-%m-%d').date()
        total_working_days = count_weekdays(start_date, end_date)

        report = {
            'missed_dates': missed_dates,
            'total_effort_hours': float(total_effort_hours),
            'broad_area_of_work_hours': [(item[0], float(item[1])) for item in broad_area_of_work_hours],
            'less_than_8_hours_dates': less_than_8_hours_dates,
            'missing_files_tasks': [(task[0], task[1]) for task in missing_files_tasks],
            'effort_towards_hours': [(item[0], float(item[1])) for item in effort_towards_hours],
            'total_working_days': total_working_days
        }
    else:
        report = None

    conn.close()

    if task_date:
        date_info = f"on {task_date}"
    else:
        date_info = f"from {from_date} to {to_date}"

    task_report = f"""
    <html>
        <head>
            <style>
                table {{
                    width: 1200px;
                    border-collapse: collapse;
                    table-layout: fixed;
                    word-wrap: break-word;
                }}
                th, td {{
                    border: 1px solid black;
                    padding: 8px;
                    text-align: left;
                }}
                th {{
                    background-color: #f2f2f2;
                }}
                .break-word {{
                    word-wrap: break-word;
                }}
            </style>
        </head>
        <body>
            <h3>Task Report for {user_id} {date_info}</h3>
            <p><strong>Employee Email:</strong> {user_email}</p>
            <p><strong>Total Effort Hours:</strong> {total_effort_hours:.2f} hours</p>
    """

    for date, tasks in tasks_by_date.items():
        daily_effort_hours = sum(task['effort_hours'] + task['effort_minutes'] / 60 for task in tasks)
        task_report += f"""
            <h4>Tasks for {date}</h4>
            <p><strong>Total Effort Hours:</strong> {daily_effort_hours:.2f} hours</p>
            <table>
                <thead>
                    <tr>
                        <th>Area of Effort</th>
                        <th>Effort (hours)</th>
                        <th>Effort Towards</th>
                        <th>Time Log Type</th>
                        <th>Output File</th>
                        <th>Output Location</th>
                        {f'<th>Manager Note</th><th>Broad Area of Work</th><th>Reviewer Note</th>' if role != 'employee' else ''}
                    </tr>
                </thead>
                <tbody>
        """
        for task in tasks:
            output_file_link = f'<a href="{task["output_file"]}" target="_blank">{task["output_file"]}</a>' if task['output_file'] else 'No output file'
            output_location_link = f'<a href="{task["output_location"]}" target="_blank">{task["output_location"]}</a>' if task['output_location'] else 'No output location'
            task_report += f"""
                    <tr>
                        <td>{task['area_of_effort']}</td>
                        <td>{task['effort_hours']}h {task['effort_minutes']}m</td>
                        <td>{task['effort_towards']}</td>
                        <td>{task['time_log_type']}</td>
                        <td class="break-word">{output_file_link}</td>
                        <td class="break-word">{output_location_link}</td>
                        {'<td>' + (task['manager_note'] or 'No manager note') + '</td><td>' + (task['broad_area_of_work'] or 'No broad area of work') + '</td><td>' + (task['reviewer_note'] or 'No reviewer note') + '</td>' if role != 'employee' else ''}
                    </tr>
            """
        task_report += """
                </tbody>
            </table>
        """

    if report:
        task_report += f"""
            <h4>Total Working Days</h4>
            <p>{report['total_working_days']} days</p>
            <h4>Missed TED Dates</h4>
            <ul>
                {''.join(f"<li>{date}</li>" for date in report['missed_dates'])}
            </ul>
            <h4>Broad Area of Work and Time Effort Hours</h4>
            <ul>
                {''.join(f"<li>{item[0]}: {item[1]:.2f} hours</li>" for item in report['broad_area_of_work_hours'])}
            </ul>
            <h4>Less than 8 hours TED Dates</h4>
            <ul>
                {''.join(f"<li>{date}</li>" for date in report['less_than_8_hours_dates'])}
            </ul>
            <h4>No Files of TED Link Missing Dates</h4>
            <ul>
        """
        for task in report['missing_files_tasks']:
            date, area_of_effort = task
            task_report += f"""
                <li>
                    <strong>{date}:</strong> {area_of_effort}
                </li>
            """
        task_report += f"""
            </ul>
            <h4>Effort Towards and Time Effort Hours</h4>
            <ul>
                {''.join(f"<li>{item[0]}: {item[1]:.2f} hours</li>" for item in report['effort_towards_hours'])}
            </ul>
        """

    task_report += """
        </body>
    </html>
    """

    # Create a directory for the employee if it doesn't exist
    employee_dir = os.path.join('reports', user_id)
    os.makedirs(employee_dir, exist_ok=True)

    # Save report to an HTML file
    html_filename = os.path.join(employee_dir, f"report_{user_id}_{date_info.replace(' ', '_').replace(':', '-')}.html")
    pdf_filename = os.path.join(employee_dir, f"report_{user_id}_{date_info.replace(' ', '_').replace(':', '-')}.pdf")
    excel_filename = os.path.join(employee_dir, f"report_{user_id}_{date_info.replace(' ', '_').replace(':', '-')}.xlsx")

    with open(html_filename, 'w') as file:
        file.write(task_report)

    # Convert HTML to PDF
    pdfkit.from_file(html_filename, pdf_filename, configuration=pdfkit_config)

    # Generate Excel report
    excel_data = []
    for date, tasks in tasks_by_date.items():
        for task in tasks:
            task_data = {
                'Task Date': date,
                'Area of Effort': task['area_of_effort'],
                'Effort (hours)': f"{task['effort_hours']}h {task['effort_minutes']}m",
                'Effort Towards': task['effort_towards'],
                'Time Log Type': task['time_log_type'],
                'Output File': task['output_file'],
                'Output Location': task['output_location']
            }
            if role != 'employee':
                task_data.update({
                    'Manager Note': task['manager_note'],
                    'Broad Area of Work': task['broad_area_of_work'],
                    'Reviewer Note': task['reviewer_note']
                })
            excel_data.append(task_data)

    # Create the DataFrame
    df_tasks = pd.DataFrame(excel_data)

    # Create a summary DataFrame
    if role == 'employee':
        summary_data = [
            {'Section': 'Employee Email', 'Details': user_email},
            {'Section': 'Total Effort Hours', 'Details': f"{total_effort_hours:.2f} hours"}
        ]
    else:
        summary_data = [
            {'Section': 'Employee Email', 'Details': user_email},
            {'Section': 'Total Effort Hours', 'Details': f"{total_effort_hours:.2f} hours"},
            {'Section': 'Total Working Days', 'Details': report['total_working_days'] if report else ''},
            {'Section': 'Missed TED Dates', 'Details': ', '.join(report['missed_dates']) if report else ''},
            {'Section': 'Broad Area of Work and Time Effort Hours', 'Details': ', '.join([f"{item[0]}: {item[1]:.2f} hours" for item in report['broad_area_of_work_hours']]) if report else ''},
            {'Section': 'Less than 8 hours TED Dates', 'Details': ', '.join(report['less_than_8_hours_dates']) if report else ''},
            {'Section': 'No Files of TED Link Missing Dates', 'Details': ', '.join([f"{task[0]}: {task[1]}" for task in report['missing_files_tasks']]) if report else ''},
            {'Section': 'Effort Towards and Time Effort Hours', 'Details': ', '.join([f"{item[0]}: {item[1]:.2f} hours" for item in report['effort_towards_hours']]) if report else ''}
        ]
    df_summary = pd.DataFrame(summary_data)

    # Write both DataFrames to Excel
    with pd.ExcelWriter(excel_filename) as writer:
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        df_tasks.to_excel(writer, sheet_name='Tasks', index=False)
 
    # Open the Excel file to adjust formatting
    workbook = load_workbook(excel_filename)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                    
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)
    workbook.save(excel_filename)

    try:
        msg = MIMEMultipart()
        msg['From'] = user_email
        msg['To'] = ADMIN_EMAIL
        msg['Subject'] = f'Task Report for {user_id} {date_info}'

        with open(html_filename, 'r') as file:
            msg.attach(MIMEText(file.read(), 'html'))

        attachment = MIMEBase('application', 'octet-stream')
        with open(pdf_filename, 'rb') as file:
            attachment.set_payload(file.read())
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', f'attachment; filename={pdf_filename}')
        msg.attach(attachment)

        attachment = MIMEBase('application', 'octet-stream')
        with open(excel_filename, 'rb') as file:
            attachment.set_payload(file.read())
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', f'attachment; filename={excel_filename}')
        msg.attach(attachment)

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(user_email, ADMIN_EMAIL, msg.as_string())

        return jsonify({'message': 'Report sent and saved successfully'}), 200
    except Exception as e:
        return jsonify({'message': f'Failed to send report: {e}'}), 500

# Payroll Management System routes

# Add employee route
@app.route('/add_employee', methods=['POST'])
def add_employee():
    data = request.json
    name = data['name']
    email = data['email']
    department = data['department']

    conn = sqlite3.connect('ted.db')
    cursor = conn.cursor()
    try:
        cursor.execute('INSERT INTO employees (name, email, department) VALUES (?, ?, ?)', (name, email, department))
        conn.commit()
        return jsonify({'message': 'Employee added successfully'}), 201
    except sqlite3.IntegrityError:
        return jsonify({'message': 'Email already exists'}), 400
    finally:
        conn.close()

# Add payroll record route
@app.route('/add_payroll', methods=['POST'])
def add_payroll():
    data = request.json
    employee_id = data['employee_id']
    period = data['period']
    salary = data['salary']
    bonus = data.get('bonus', 0)
    deductions = data.get('deductions', 0)
    tax = data['tax']
    net_salary = salary + bonus - deductions - tax

    conn = sqlite3.connect('ted.db')
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO payroll (employee_id, period, salary, bonus, deductions, tax, net_salary)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (employee_id, period, salary, bonus, deductions, tax, net_salary))
        conn.commit()
        return jsonify({'message': 'Payroll record added successfully'}), 201
    except sqlite3.IntegrityError:
        return jsonify({'message': 'Failed to add payroll record'}), 400
    finally:
        conn.close()

# Get payroll records route
@app.route('/get_payroll_records', methods=['GET'])
def get_payroll_records():
    employee_id = request.args.get('employee_id')
    conn = sqlite3.connect('ted.db')
    cursor = conn.cursor()
    query = 'SELECT * FROM payroll WHERE employee_id = ?'
    cursor.execute(query, (employee_id,))
    payroll_records = cursor.fetchall()
    conn.close()
    payroll_list = [
        {
            'id': record[0],
            'employee_id': record[1],
            'period': record[2],
            'salary': record[3],
            'bonus': record[4],
            'deductions': record[5],
            'tax': record[6],
            'net_salary': record[7]
        } for record in payroll_records
    ]
    return jsonify(payroll_list), 200

# Serve the React app
@app.route('/')
def serve_home():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory(app.static_folder, path)

if __name__ == '__main__':
    app.run(debug=True)
